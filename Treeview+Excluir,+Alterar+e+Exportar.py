#Cadastrar itens
from tkinter import *
from tkinter import ttk
from tkinter import messagebox


#tk - Biblioteca do tkinter
#Tk - Janela / Tela
janela = Tk()

#Defeni um tamanho fixo para a tela
janela.geometry("900x400")

#Altera o título da tela
janela.title("Treeview exemplo 2")

#grid - Divide a tela em partes / grades
#sticky - Usamos para preencher os espaços vazios nas laterais
#sticky - NSEW (Norte, Sul, Leste e Oeste)
id = Label(text="ID: ",
           font="Arial 12").grid(row=1, column=0, sticky="E")

#Campo digitavel
exibirID = Entry(font="Arial 12")
exibirID.grid(row=1, column=1, sticky="W")

#----------------------------

nome = Label(text="Nome: ",
           font="Arial 12").grid(row=1, column=2, sticky="E")

#Campo digitavel
exibirNome = Entry(font="Arial 12")
exibirNome.grid(row=1, column=3, sticky="W")

#----------------------------

idade = Label(text="Idade: ",
           font="Arial 12").grid(row=1, column=4, sticky="E")

#Campo digitavel
exibirIdade = Entry(font="Arial 12")
exibirIdade.grid(row=1, column=5, sticky="W")

#----------------------------

sexo = Label(text="Sexo: ",
           font="Arial 12").grid(row=1, column=6, sticky="E")

#Campo digitavel
exibirSexo = Entry(font="Arial 12")
exibirSexo.grid(row=1, column=7, sticky="W")

def addItemTreeView():

    #if - Se
    #elif - Senão Se
    #else - Senão
    if exibirID.get() == "":
        messagebox.showinfo("Atenção!", "Digite um ID")

    elif exibirNome.get() == "":
        messagebox.showinfo("Atenção!", "Digite o Nome")

    elif exibirIdade.get() == "":
        messagebox.showinfo("Atenção!", "Digite a Idade")

    elif exibirSexo.get() == "":
        messagebox.showinfo("Atenção!", "Digite o Sexo")

    else:

        #Cadastrando o item
        treeviewDados.insert("", "end",
                             values=(str(exibirID.get()),
                                     str(exibirNome.get()),
                                     str(exibirIdade.get()),
                                     str(exibirSexo.get())))

        messagebox.showinfo(title="Atenção!", message="Dados cadastrados com sucesso!")

        # Chamando a nossa função que conta a quantidade de linhas que tem na Treeview
        contarNumeroLinhas()

        #Limpar os itens
        exibirID.delete(0, "end")
        exibirNome.delete(0, "end")
        exibirIdade.delete(0, "end")
        exibirSexo.delete(0, "end")


botaoAdicionar = Button(text="Cadastrar", font="Arial 20",
                        command= addItemTreeView)

#columnspan - É a quantidade de colunas que o item vai oculpar no grid
botaoAdicionar.grid(row=2, column=0, columnspan=2, sticky="NSEW", padx=5, pady=5)

#clam, alt, default, classic
stilo = ttk.Style()
stilo.theme_use("alt")
#stilo.configure(".")

treeviewDados = ttk.Treeview(janela, columns=(1, 2, 3, 4), show="headings")

#Populando o titulo
treeviewDados.column("1", anchor=CENTER)
treeviewDados.heading("1", text="ID")

treeviewDados.column("2", anchor=CENTER)
treeviewDados.heading("2", text="Nome")

treeviewDados.column("3", anchor=CENTER)
treeviewDados.heading("3", text="Idade")

treeviewDados.column("4", anchor=CENTER)
treeviewDados.heading("4", text="Sexo")

treeviewDados.insert("", "end", text="1", values=("1", "Allan", 29, "Masculino"))
treeviewDados.insert("", "end", text="2", values=("2", "Ana", 41, "Feminino"))
treeviewDados.insert("", "end", text="3", values=("3", "Berenice", 50, "Feminino"))
treeviewDados.insert("", "end", text="4", values=("4", "Roger", 21, "Masculino"))
treeviewDados.insert("", "end", text="5", values=("5", "Pedro", 45, "Masculino"))

#columnspan - É a quantidade de colunas que o item vai oculpar no grid
treeviewDados.grid(row=3, column=0, columnspan=8, sticky="NSEW")

#Função para deletar o item
def deletarItemTreeView():

    selecionarItens = treeviewDados.selection()

    #for - para
    for item in selecionarItens:

        #Deletando o item selecionado
        treeviewDados.delete(item)

        messagebox.showinfo(title="Atenção!", message="Nome deletado com sucesso!")

        # Chamando a nossa função que conta a quantidade de linhas que tem na Treeview
        contarNumeroLinhas()

        #Limpar os itens / caixa de texto
        exibirID.delete(0, "end")
        exibirNome.delete(0, "end")
        exibirIdade.delete(0, "end")
        exibirSexo.delete(0, "end")

botaoExcluir = Button(text="Excluir", font="Arial 20",
                        command= deletarItemTreeView)

#columnspan - É a quantidade de colunas que o item vai oculpar no grid
botaoExcluir.grid(row=2, column=2, columnspan=2, sticky="NSEW", padx=5, pady=5)


def alterarItemTreeView():

    itemSelecionado = treeviewDados.selection()

    # if - Se
    # elif - Senão Se
    # else - Senão
    if exibirID.get() == "":
        messagebox.showinfo("Atenção!", "Digite um ID")

    elif exibirNome.get() == "":
        messagebox.showinfo("Atenção!", "Digite o Nome")

    elif exibirIdade.get() == "":
        messagebox.showinfo("Atenção!", "Digite a Idade")

    elif exibirSexo.get() == "":
        messagebox.showinfo("Atenção!", "Digite o Sexo")

    else:

        #Alterar o item
        treeviewDados.item(itemSelecionado,
                           values=(str(exibirID.get()),
                                   str(exibirNome.get()),
                                   str(exibirIdade.get()),
                                   str(exibirSexo.get())))

        messagebox.showinfo(title="Atenção!", message="Item alterado com sucesso!")

        # Chamando a nossa função que conta a quantidade de linhas que tem na Treeview
        contarNumeroLinhas()

        # Limpar os itens
        exibirID.delete(0, "end")
        exibirNome.delete(0, "end")
        exibirIdade.delete(0, "end")
        exibirSexo.delete(0, "end")


botaoAlterar = Button(text="Alterar", font="Arial 20",
                        command= alterarItemTreeView)

#columnspan - É a quantidade de colunas que o item vai oculpar no grid
botaoAlterar.grid(row=2, column=4, columnspan=2, sticky="NSEW", padx=5, pady=5)


#load_workbook
from openpyxl import load_workbook
def exportarParaExcel():

    #Selecionando o arquivo
    workbook = load_workbook(filename="A:\\TKinter\\Primeiros Passos\\Tratamento_Dados.xlsx")

    #Selecionando a sheet / planilha
    sheet = workbook["Produtos"]

    #Deletando a linhas
    sheet.delete_rows(idx=1, amount=30000)

    #for - para
    for numeroLinha in treeviewDados.get_children():

        #Pegando os dados da linha que estiver passando
        linha = treeviewDados.item(numeroLinha)["values"]

        #Passando os dados da linha para a planilha do Excel
        sheet.append(linha)

    #Salva a planilha com todas as informações do TreeView
    workbook.save(filename="A:\\TKinter\\Primeiros Passos\\Dados_Exportados_Treeview.xlsx")
    messagebox.showinfo(title="Atenção!", message="Dados exportados com sucesso!")

botaoExportar = Button(text="Exportar", font="Arial 20",
                        command= exportarParaExcel)

#columnspan - É a quantidade de colunas que o item vai oculpar no grid
botaoExportar.grid(row=2, column=6, columnspan=2, sticky="NSEW", padx=5, pady=5)

labelNumeroLinhas = Label(text="Linhas: ",
                          font="Arial 20")

def contarNumeroLinhas(item=""):

    numero = 0

    linhas = treeviewDados.get_children(item)

    #for - para
    for linha in linhas:

        #Somo linha por linha para contar quantas tem na Treeview
        #numero = numero + 1
        numero += 1

    #Imprimo o total de linhas no Label
    labelNumeroLinhas.config(text="Linhas: " + str(numero))

#columnspan - É a quantidade de colunas que o item vai oculpar no grid
labelNumeroLinhas.grid(row=4, column=0, columnspan=8, sticky="W")

#Chamando a nossa função que conta a quantidade de linhas que tem na Treeview
contarNumeroLinhas()

#mainloop - No Tkinter uma janela funciona como um loop infinito
#A janela que o Python mostra na verdade é um programa em funcionamento
janela.mainloop()